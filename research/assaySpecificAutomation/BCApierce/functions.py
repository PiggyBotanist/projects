import os
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats
import openpyxl as xl

# load values from excel sheet.
def load_value(location, shape, excelFile):
    # location = left top most position (row,column)
    # shape = shape of the values (row, column)
    # excelFile = the file you want to load from (use openpyxl)
    output = np.zeros(shape)
    for i in range(0, shape[0]):
        for j in range(0, shape[1]):
            output[i][j] = excelFile.cell(row = location[0] + i, column = location[1] + j).value
    return output

# load single value from excel sheet.
def load_single_value(location, excelFile):
    return excelFile.cell(row = location[0], column = location[1]).value

# output values to excel sheet.
def output_value(location, shape, values, excelFile):
    for i in range(0, shape[0]):
        for j in range(0, shape[1]):
            if values[i][j] <= 0:
                excelFile.cell(row=location[0] + i, column=location[1] + j).value = np.nan
            else:
                excelFile.cell(row = location[0] + i, column=location[1] + j).value = values[i][j]

# output single value to excelsheet.
def output_single_value(location, value, excelFile):
    excelFile.cell(row = location[0], column = location[1]).value = value

# find corrector to od value.
def find_corrector(standard_curve_od_matrix, repeats, excelFile):
    if standard_curve_od_matrix[0][0] < standard_curve_od_matrix[7][0]:
        return np.sum(load_value((2, 2), (1, repeats), excelFile)) / repeats
    else:
        return np.sum(load_value((9, 2), (1, repeats), excelFile)) / repeats

# remove any negatives or zeros
def nullify_negatives(matrix):
    for i in range(0,len(matrix)):
        for j in range(0, len(matrix[0])):
            if matrix[i][j] <= 0:
                matrix[i][j] = 0
    return matrix

# calculate protein concentration
def extrapolate(matrix, slope, intercept, dilution_factor):
    for i in range(0,len(matrix)):
        for j in range(0, len(matrix[0])):
            if matrix[i][j] != 0:
                matrix[i][j] = (matrix[i][j] - intercept) * dilution_factor / slope
    return matrix
